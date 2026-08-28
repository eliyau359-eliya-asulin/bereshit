"""Bulk product import from Excel (.xlsx) or CSV. Two-step, never
one-shot: validate_rows() only ever reads the database (never writes),
producing a full per-row report the admin reviews in the UI; apply_import()
is a separate call that actually writes, and only after the admin has
seen and confirmed the preview. A single bad row never corrupts the
catalog — the whole apply runs in one MongoDB transaction, so a failure
partway through leaves nothing written at all.

Column names intentionally match the product's actual current fields —
name/description/material are single-language today (no separate
Hebrew/English variants exist in the schema yet), so the import format
mirrors that rather than inventing bilingual columns the rest of the
system doesn't support.
"""
import csv
import io
from datetime import datetime, timezone

from openpyxl import load_workbook
from pymongo import ReturnDocument

from backend.db.mongo import get_db, get_client
from backend.models.schemas import ValidationError

COLUMNS = ["SKU", "Barcode", "Name", "Category", "Price", "OldPrice", "Stock", "Threshold",
           "ShortDescription", "FullDescription", "Material", "Dimensions", "ImageURL"]
REQUIRED_COLUMNS = {"SKU", "Name", "Category", "Price", "Stock"}
MAX_ROWS = 2000


def build_template_csv():
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(COLUMNS)
    writer.writerow([
        "BJ-EXAMPLE-1", "7290000000001", 'גביע קידוש לדוגמה', "השם המדויק של קטגוריה קיימת",
        "199", "249", "10", "5", "תיאור קצר לדוגמה", "תיאור מלא לדוגמה", "כסף סטרלינג 925",
        'גובה 15 ס"מ', "",
    ])
    return buf.getvalue().encode("utf-8-sig")  # BOM so Excel opens Hebrew text correctly, not as mojibake


def parse_import_file(raw_bytes, filename):
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        rows = _parse_xlsx(raw_bytes)
    elif name.endswith(".csv"):
        rows = _parse_csv(raw_bytes)
    else:
        raise ValidationError("סוג קובץ לא נתמך — יש להעלות קובץ .xlsx או .csv")

    if not rows:
        raise ValidationError("הקובץ ריק או שלא נמצאו בו שורות נתונים")
    if len(rows) > MAX_ROWS:
        raise ValidationError(f"הקובץ מכיל יותר מדי שורות (מקסימום {MAX_ROWS})")

    header_set = set(rows[0].keys())
    missing = REQUIRED_COLUMNS - header_set
    if missing:
        raise ValidationError("חסרות עמודות חובה בקובץ: " + ", ".join(sorted(missing)))

    return rows


def _parse_xlsx(raw_bytes):
    try:
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception:
        raise ValidationError("לא ניתן לקרוא את קובץ ה-Excel — ודא שזהו קובץ .xlsx תקין")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    out = []
    for r in all_rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})
    return out


def _parse_csv(raw_bytes):
    try:
        text = raw_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValidationError("לא ניתן לקרוא את קובץ ה-CSV — יש לשמור אותו בקידוד UTF-8")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader if any((v or "").strip() for v in row.values())]


def _to_number(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(val, default=None):
    n = _to_number(val)
    if n is None:
        return default
    return int(n)


def _cell(row, key):
    v = row.get(key)
    if v is None:
        return ""
    # An .xlsx cell holding a digit-only value (a SKU or barcode typed
    # without first formatting the column as text — the common case)
    # comes back from openpyxl as a Python float, e.g. 7290000000001.0.
    # str()'ing that directly would bake a literal ".0" onto a real-world
    # barcode/SKU, so a whole-number float is rendered without the
    # trailing decimal — same as what the admin actually typed.
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def validate_rows(raw_rows):
    db = get_db()
    categories_by_label = {c["label"].strip(): c for c in db.categories.find()}
    existing_by_sku = {p["sku"]: p for p in db.products.find()}
    existing_barcodes = {p["barcode"] for p in db.products.find({"barcode": {"$nin": [None, ""]}}, {"barcode": 1})}

    results = []
    seen_skus = set()
    seen_barcodes = set()

    for idx, raw in enumerate(raw_rows, start=2):  # row 1 is the header row
        errors = []
        sku = _cell(raw, "SKU")
        if not sku:
            errors.append('מק"ט (SKU) חסר')
        elif sku in seen_skus:
            errors.append(f'מק"ט \'{sku}\' מופיע יותר מפעם אחת בקובץ')
        seen_skus.add(sku)

        name = _cell(raw, "Name")
        if not name:
            errors.append("שם המוצר חסר")

        category_label = _cell(raw, "Category")
        category = categories_by_label.get(category_label)
        if not category_label:
            errors.append("קטגוריה חסרה")
        elif not category:
            errors.append(f"קטגוריה '{category_label}' אינה קיימת — יש ליצור אותה קודם או לתקן את השם")

        price = _to_number(raw.get("Price"))
        if price is None or price < 0:
            errors.append("מחיר חסר או לא תקין")

        old_price_raw = _cell(raw, "OldPrice")
        old_price = _to_number(old_price_raw) if old_price_raw else None
        if old_price_raw and old_price is None:
            errors.append("מחיר קודם אינו מספר תקין")

        stock = _to_int(raw.get("Stock"))
        if stock is None or stock < 0:
            errors.append("כמות מלאי חסרה או לא תקינה")

        threshold = _to_int(raw.get("Threshold"), default=5)
        if threshold is None or threshold < 0:
            errors.append("סף מלאי נמוך אינו תקין")

        barcode = _cell(raw, "Barcode") or None
        sku_for_lookup = sku or None
        existing = existing_by_sku.get(sku_for_lookup)
        action = "update" if existing else "create"

        if barcode:
            belongs_to_this_row_already = existing and existing.get("barcode") == barcode
            if not belongs_to_this_row_already:
                if barcode in seen_barcodes:
                    errors.append(f"ברקוד '{barcode}' מופיע יותר מפעם אחת בקובץ")
                elif barcode in existing_barcodes:
                    errors.append(f"ברקוד '{barcode}' כבר קיים במוצר אחר במערכת")
            seen_barcodes.add(barcode)

        clean = {
            "sku": sku, "name": name,
            # Categories are stored with `_id` as the key (see
            # categories_service.py) — this reads the raw collection
            # directly, not through that service's _serialize(), so it's
            # `_id` here, not `key`.
            "cat": category["_id"] if category else None,
            "catLabel": category["label"] if category else category_label,
            "price": price, "oldPrice": old_price,
            "stock": stock, "threshold": threshold,
            "short": _cell(raw, "ShortDescription"),
            "desc": _cell(raw, "FullDescription"),
            "material": _cell(raw, "Material"),
            "dim": _cell(raw, "Dimensions"),
            "image": _cell(raw, "ImageURL") or None,
        }
        # Same sparse+unique-index reasoning as products_service.create_product:
        # a blank Barcode cell must leave the key entirely absent (never an
        # explicit null), or a second imported row without a barcode collides
        # with the first. A blank cell on an UPDATE row must also not appear
        # in the patch at all, so it doesn't silently wipe out a barcode the
        # product already has just because this re-upload's row omitted it.
        if barcode:
            clean["barcode"] = barcode

        results.append({
            "row": idx, "sku": sku, "name": name, "action": action,
            "valid": not errors, "errors": errors, "data": clean,
        })

    valid_rows = [r for r in results if r["valid"]]
    return {
        "rows": results,
        "summary": {
            "total": len(results),
            "valid": len(valid_rows),
            "invalid": len(results) - len(valid_rows),
            "toCreate": sum(1 for r in valid_rows if r["action"] == "create"),
            "toUpdate": sum(1 for r in valid_rows if r["action"] == "update"),
        },
    }


def apply_import(rows, actor=None):
    """Re-validates server-side (never trusts that the rows the client
    sends back from a preview still match the current DB state — a
    category or another product could have changed in between) and
    writes everything in one transaction. Only rows the caller marked
    valid are ever attempted; anything else is skipped and reported."""
    valid_rows = [r for r in rows if r.get("valid") and r.get("data")]
    if not valid_rows:
        raise ValidationError("אין שורות תקינות לייבוא")

    db = get_db()
    client = get_client()
    result_holder = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    def _run(session):
        created = updated = skipped = 0
        errors = []

        # One query for every SKU this batch touches, instead of a find_one
        # per row — the same batching validate_rows already does, now
        # reused here too rather than re-querying per row inside the
        # transaction (which also held the transaction open longer under
        # a big import, and the original find_one(sort=[("_id",-1)]) per
        # new row was a race: two concurrent imports could both read the
        # same "last id" and try to insert the same _id. Reading it once
        # per transaction and incrementing a local counter closes that —
        # any real interleaving now surfaces as a MongoDB write conflict,
        # which with_transaction retries automatically with fresh data.
        skus = [r["data"]["sku"] for r in valid_rows]
        existing_by_sku = {p["sku"]: p for p in db.products.find({"sku": {"$in": skus}}, session=session)}
        last = db.products.find_one(sort=[("_id", -1)], session=session)
        next_id = (last["_id"] + 1) if last else 1

        for r in valid_rows:
            data = r["data"]
            sku = data["sku"]
            if not data.get("cat"):
                skipped += 1
                errors.append({"row": r["row"], "sku": sku, "error": "קטגוריה אינה תקינה עוד — דולג"})
                continue

            existing = existing_by_sku.get(sku)
            if existing:
                before_stock = existing.get("stock", 0)
                patch = {k: v for k, v in data.items() if k != "sku"}
                db.products.update_one({"_id": existing["_id"]}, {"$set": patch}, session=session)
                if patch.get("stock") is not None and patch["stock"] != before_stock:
                    db.inventory_log.insert_one({
                        "productId": existing["_id"],
                        "productName": data["name"],
                        "previousStock": before_stock,
                        "newStock": patch["stock"],
                        "delta": patch["stock"] - before_stock,
                        "reason": "ייבוא מקובץ Excel/CSV",
                        "actor": actor,
                        "at": datetime.now(timezone.utc).isoformat(),
                    }, session=session)
                updated += 1
            else:
                doc = dict(data)
                doc["_id"] = next_id
                next_id += 1
                doc.setdefault("oldPrice", None)
                doc.setdefault("badge", None)
                doc.setdefault("status", "active")
                doc.setdefault("sold", 0)
                doc.setdefault("image", None)
                doc.setdefault("thumbnail", None)
                # NOT barcode — see products_service.create_product's
                # identical comment: the field must stay genuinely absent
                # when blank, never an explicit null, or a second
                # barcode-less product collides on the unique sparse index.
                db.products.insert_one(doc, session=session)
                created += 1

        result_holder.update(created=created, updated=updated, skipped=skipped, errors=errors)

    with client.start_session() as session:
        session.with_transaction(_run)

    return result_holder
